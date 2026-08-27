"""확정된 학생별 문항 점수로 단순한 엑셀 성적표를 만든다."""

import io

from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Alignment, Font

from app.services.feedback_builder import FeedbackContext


FORMULA_PREFIXES = ("=", "+", "-", "@", "\t", "\r")


def _excel_text(value: str) -> str:
    cleaned = ILLEGAL_CHARACTERS_RE.sub("", value)
    if cleaned.startswith(FORMULA_PREFIXES):
        return "'" + cleaned
    return cleaned


def _layout(context: FeedbackContext) -> list[tuple[int, int]]:
    return [(item.item_no, item.max_points) for item in context.items]


def _validate_contexts(contexts: list[FeedbackContext]) -> list[tuple[int, int]]:
    if not contexts:
        return []
    reference = _layout(contexts[0])
    if len({item_no for item_no, _ in reference}) != len(reference):
        raise ValueError("성적표 문항 번호가 중복되었습니다.")
    student_numbers: set[int] = set()
    for context in contexts:
        if _layout(context) != reference:
            raise ValueError("학생별 성적표 문항 구성이 서로 다릅니다.")
        if context.student_number in student_numbers:
            raise ValueError("성적표 학생 번호가 중복되었습니다.")
        student_numbers.add(context.student_number)
        if any(not 0 <= item.score <= item.max_points for item in context.items):
            raise ValueError("성적표 문항 점수가 배점과 맞지 않습니다.")
        if sum(item.score for item in context.items) != context.total_score:
            raise ValueError("성적표 총점과 문항 점수 합계가 다릅니다.")
    return reference


def build_gradebook(contexts: list[FeedbackContext], title: str) -> bytes:
    layout = _validate_contexts(contexts)
    item_numbers = [item_no for item_no, _ in layout]
    workbook = Workbook()
    workbook.properties.title = _excel_text(title)[:255]
    sheet = workbook.active
    if sheet is None:
        raise RuntimeError("성적표 시트를 만들지 못했습니다.")
    sheet.title = "성적"
    sheet.append(
        ["번호", "이름"]
        + [f"{item_no}번" for item_no in item_numbers]
        + ["총점", "성취수준"]
    )
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    for context in contexts:
        by_item = {item.item_no: item.score for item in context.items}
        sheet.append(
            [context.student_number, _excel_text(context.student_name)]
            + [by_item[item_no] for item_no in item_numbers]
            + [context.total_score, f"{context.level}수준"]
        )

    sheet.column_dimensions["B"].width = 14
    sheet.freeze_panes = "C2"
    sheet.auto_filter.ref = sheet.dimensions
    _add_item_summary(workbook, contexts, layout)

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _add_item_summary(
    workbook: Workbook,
    contexts: list[FeedbackContext],
    layout: list[tuple[int, int]],
) -> None:
    sheet = workbook.create_sheet("문항별 통계")
    sheet.append(["문항", "배점", "평균", "평균 득점률"])
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    for item_no, max_points in layout:
        scores = [
            next(item.score for item in context.items if item.item_no == item_no)
            for context in contexts
        ]
        average = sum(scores) / len(scores)
        rate = average / max_points if max_points else 0.0
        sheet.append(
            [f"{item_no}번", max_points, round(average, 2), round(rate, 4)]
        )
        sheet.cell(sheet.max_row, 4).number_format = "0.0%"

    sheet.column_dimensions["A"].width = 10
    sheet.freeze_panes = "A2"
