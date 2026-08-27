import io
from dataclasses import replace

import pytest
from openpyxl import load_workbook

from app.services.export_excel import build_gradebook
from tests.test_feedback_generator import _context as make_context


def _contexts():
    first = make_context()
    second_items = [
        first.items[0],
        replace(first.items[1], score=2),
    ]
    second = replace(
        make_context(),
        student_name="박균형",
        student_number=2,
        total_score=4,
        level="3",
        items=second_items,
    )
    return [first, second]


def test_workbook_has_header_and_student_rows():
    data = build_gradebook(_contexts(), title="수학 논술형")
    sheet = load_workbook(io.BytesIO(data)).active

    header = [cell.value for cell in sheet[1]]
    assert header == ["번호", "이름", "1번", "2번", "총점", "성취수준"]
    assert sheet.max_row == 3


def test_item_scores_totals_and_level_are_written():
    data = build_gradebook(_contexts(), title="수학 논술형")
    sheet = load_workbook(io.BytesIO(data)).active
    row = [cell.value for cell in sheet[2]]

    assert row == [1, "김미래", 2, 1, 3, "2수준"]


def test_summary_sheet_reports_item_averages():
    data = build_gradebook(_contexts(), title="수학 논술형")
    workbook = load_workbook(io.BytesIO(data))
    sheet = workbook["문항별 통계"]

    assert [cell.value for cell in sheet[1]] == [
        "문항",
        "배점",
        "평균",
        "평균 득점률",
    ]
    assert sheet.cell(2, 3).value == 2
    assert sheet.cell(3, 3).value == 1.5
    assert sheet.cell(3, 4).value == 0.75


def test_empty_contexts_still_produce_headers():
    data = build_gradebook([], title="빈 평가")
    workbook = load_workbook(io.BytesIO(data))

    assert workbook.active.max_row == 1
    assert workbook["문항별 통계"].max_row == 1


def test_spreadsheet_formula_prefix_is_escaped():
    dangerous = replace(make_context(), student_name="=1+1")
    data = build_gradebook([dangerous], title="평가")
    cell = load_workbook(io.BytesIO(data)).active.cell(2, 2)

    assert cell.data_type != "f"
    assert cell.value == "'=1+1"


def test_inconsistent_item_layout_is_rejected():
    second = replace(make_context(), items=make_context().items[:1])
    with pytest.raises(ValueError, match="문항"):
        build_gradebook([make_context(), second], title="평가")
