import io

import pytest
from openpyxl import load_workbook

from app.api import feedback as feedback_api
from app.api import settings as settings_api
from app.models.app_setting import (
    DATA_POLICY_WORDING_TEXT,
    DATA_POLICY_WORDING_VERSION,
    KEY_DATA_POLICY_ACK,
    KEY_LLM_MODEL,
    KEY_LLM_MODEL_KEY_FINGERPRINT,
    AppSetting,
    DataPolicyAcknowledgement,
)
from app.models.feedback import Feedback
from app.models.grading import ItemScore
from app.models.rubric import RubricDraft
from app.providers.credentials import CredentialStore
from app.services.confirmation import confirm_score
from app.services.feedback_generator import GeneratedFeedback
from tests.fakes import FakeKeyring


RUBRIC_CONTENT = {
    "assessment": {"title": "가", "subject": "수학", "grade": 6, "total_points": 4},
    "achievement_standards": [
        {
            "id": "AS1",
            "item_range": [1, 2],
            "core_standard": "대칭을 이해한다.",
            "levels": {"1": "시도", "2": "가능", "3": "정확"},
        }
    ],
    "items": [
        {
            "item_no": 1,
            "title": "대칭축 찾기",
            "points": 2,
            "standard_id": "AS1",
            "type": "closed_short",
            "blanks": [{"key": "1", "answers": ["선대칭"], "aliases": []}],
            "scoring": [
                {"score": 2, "condition": "all_correct", "criterion": "둘 다 정확"},
                {"score": 0, "condition": "none_correct", "criterion": "모두 틀림"},
            ],
            "example_answer": "1 선대칭",
        },
        {
            "item_no": 2,
            "title": "점대칭 도형 그리기",
            "points": 2,
            "standard_id": "AS1",
            "type": "drawing",
            "scoring": [
                {"score": 2, "condition": "manual", "criterion": "정확하게 완성"},
                {"score": 0, "condition": "manual", "criterion": "그리지 못함"},
            ],
            "example_answer": "예시 그림",
        },
    ],
    "level_cutoffs": {"3": 4, "2": 2, "1": 0},
}


@pytest.fixture
def ready(graded_setup, db_session, client, tmp_path):
    assessment = graded_setup["assessment"]
    assessment.level_cutoffs = {"3": 4, "2": 2, "1": 0}
    assessment.status = "confirmed"
    db_session.add(
        RubricDraft(
            assessment_id=assessment.id,
            confirmed=True,
            content=RUBRIC_CONTENT,
        )
    )
    store = CredentialStore(tmp_path / "credential", FakeKeyring())
    store.set_api_key("test-api-key")
    db_session.add_all(
        [
            AppSetting(key=KEY_LLM_MODEL, value="models/test-model"),
            AppSetting(
                key=KEY_LLM_MODEL_KEY_FINGERPRINT,
                value=settings_api._key_fingerprint("test-api-key"),
            ),
            AppSetting(key=KEY_DATA_POLICY_ACK, value="true"),
            DataPolicyAcknowledgement(
                wording_version=DATA_POLICY_WORDING_VERSION,
                wording_text=DATA_POLICY_WORDING_TEXT,
                acknowledged=True,
            ),
        ]
    )
    db_session.commit()
    client.app.dependency_overrides[feedback_api.get_credential_store] = lambda: store
    return graded_setup


@pytest.fixture
def stub_generation(monkeypatch):
    monkeypatch.setattr(
        feedback_api,
        "build_feedback_provider",
        lambda *_args, **_kwargs: object(),
    )
    monkeypatch.setattr(
        feedback_api,
        "generate_one",
        lambda context, *_args, **_kwargs: GeneratedFeedback(
            item_comments=[
                {
                    "item_no": item.item_no,
                    "score": item.score,
                    "max": item.max_points,
                    "comment": f"{item.item_no}번 설명",
                }
                for item in context.items
            ],
            summary="학생은 대칭을 이해하고 있어요.",
            next_step="다른 대칭 문제도 풀어 보세요.",
        ),
    )


def _confirm_all(db_session, run_id, value=2):
    for score in db_session.query(ItemScore).filter_by(run_id=run_id):
        confirm_score(
            db_session,
            score,
            new_score=value,
            source="teacher_accept",
        )


def test_generation_blocked_while_unconfirmed(ready, client, stub_generation):
    response = client.post(f"/api/feedback/runs/{ready['run'].id}/generate")
    assert response.status_code == 400
    assert "확정" in response.json()["detail"]


def test_generation_requires_current_policy_event(
    ready, client, stub_generation, db_session
):
    _confirm_all(db_session, ready["run"].id)
    event = db_session.query(DataPolicyAcknowledgement).one()
    event.wording_text = "이전 문구"
    db_session.commit()

    response = client.post(f"/api/feedback/runs/{ready['run'].id}/generate")
    assert response.status_code == 403


def test_generation_requires_model_bound_to_current_key(
    ready, client, stub_generation, db_session
):
    _confirm_all(db_session, ready["run"].id)
    binding = db_session.query(AppSetting).filter_by(
        key=KEY_LLM_MODEL_KEY_FINGERPRINT
    ).one()
    binding.value = "0" * 64
    db_session.commit()

    response = client.post(f"/api/feedback/runs/{ready['run'].id}/generate")
    assert response.status_code == 400


def test_generates_and_lists_feedback_for_each_student(
    ready, client, stub_generation, db_session
):
    run_id = ready["run"].id
    _confirm_all(db_session, run_id)

    response = client.post(f"/api/feedback/runs/{run_id}/generate")
    assert response.status_code == 200
    assert response.json() == {"generated": 2, "degraded": 0}
    body = client.get(f"/api/feedback/runs/{run_id}").json()
    assert [row["student_name"] for row in body["feedbacks"]] == [
        "김미래",
        "박균형",
    ]
    assert body["feedbacks"][0]["level"] == "3"
    assert body["feedbacks"][0]["stale"] is False


def test_html_and_excel_exports(ready, client, stub_generation, db_session):
    run_id = ready["run"].id
    _confirm_all(db_session, run_id)
    client.post(f"/api/feedback/runs/{run_id}/generate")

    html = client.get(f"/api/feedback/runs/{run_id}/export.html")
    assert html.status_code == 200
    assert "김미래" in html.text and "박균형" in html.text
    assert html.headers["cache-control"] == "no-store"
    assert "default-src 'none'" in html.text

    excel = client.get(f"/api/feedback/runs/{run_id}/export.xlsx")
    assert excel.status_code == 200
    assert "attachment" in excel.headers["content-disposition"]
    sheet = load_workbook(io.BytesIO(excel.content)).active
    assert sheet.max_row == 3
    assert sheet.cell(row=2, column=2).value == "김미래"


def test_exports_before_generation_return_404(ready, client, db_session):
    _confirm_all(db_session, ready["run"].id)
    run_id = ready["run"].id
    assert client.get(f"/api/feedback/runs/{run_id}/export.html").status_code == 404


def test_regenerating_replaces_previous(ready, client, stub_generation, db_session):
    run_id = ready["run"].id
    _confirm_all(db_session, run_id)
    client.post(f"/api/feedback/runs/{run_id}/generate")
    client.post(f"/api/feedback/runs/{run_id}/generate")

    assert db_session.query(Feedback).filter_by(run_id=run_id).count() == 2


def test_score_change_marks_feedback_stale_and_blocks_export(
    ready, client, stub_generation, db_session
):
    run_id = ready["run"].id
    _confirm_all(db_session, run_id)
    client.post(f"/api/feedback/runs/{run_id}/generate")
    score = db_session.query(ItemScore).filter_by(run_id=run_id).first()
    confirm_score(db_session, score, new_score=0, source="teacher_edit")

    rows = client.get(f"/api/feedback/runs/{run_id}").json()["feedbacks"]
    assert any(row["stale"] for row in rows)
    assert client.get(f"/api/feedback/runs/{run_id}/export.html").status_code == 409


def test_unexpected_generation_failure_keeps_previous_feedback(
    ready, client, stub_generation, db_session, monkeypatch
):
    run_id = ready["run"].id
    _confirm_all(db_session, run_id)
    client.post(f"/api/feedback/runs/{run_id}/generate")
    monkeypatch.setattr(
        feedback_api,
        "generate_one",
        lambda *_args, **_kwargs: (_ for _ in ()).throw(RuntimeError("secret")),
    )

    response = client.post(f"/api/feedback/runs/{run_id}/generate")
    assert response.status_code == 502
    assert "secret" not in response.text
    assert db_session.query(Feedback).filter_by(run_id=run_id).count() == 2
